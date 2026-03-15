"""
Excel Export 모듈
Sheet1: Bus Summary / Sheet2: Route / Sheet3: Passenger
"""
import io
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

# 버스별 색상 팔레트
BUS_COLORS = [
    "4472C4", "ED7D31", "A9D18E", "FF0000", "7030A0",
    "00B0F0", "FFFF00", "92D050", "00B050", "FF7F50"
]

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_ROW_FILL = PatternFill(start_color="F2F7FF", end_color="F2F7FF", fill_type="solid")


def thin_border():
    thin = Side(style='thin', color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


class ExcelExporter:
    def export(self, routes: List[Dict], destination: Dict,
               summary: Optional[Dict] = None) -> bytes:
        wb = Workbook()

        self._write_summary_sheet(wb, routes, destination, summary)
        self._write_route_sheet(wb, routes)
        self._write_passenger_sheet(wb, routes)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _write_summary_sheet(self, wb, routes, destination, summary):
        ws = wb.active
        ws.title = "Bus Summary"

        # 제목
        ws.merge_cells("A1:G1")
        ws["A1"] = "🚌 버스 노선 최적화 결과"
        ws["A1"].font = Font(size=16, bold=True, color="1F3864")
        ws["A1"].alignment = Alignment(horizontal="center")

        # 행사장 정보
        ws["A2"] = f"행사장: {destination.get('address', '')}  |  도착 목표시간: {destination.get('arrival_time', '')}"
        ws["A2"].font = Font(size=11, italic=True, color="595959")
        ws.merge_cells("A2:G2")

        # 요약 정보
        if summary:
            ws["A3"] = (f"총 승객: {summary.get('total_passengers', 0)}명  |  "
                        f"버스 {summary.get('total_buses', 0)}대 운행  |  "
                        f"총 이동: {summary.get('total_duration_min', 0)}분")
            ws["A3"].font = Font(size=10, color="595959")
            ws.merge_cells("A3:G3")

        ws.append([])  # 빈 행

        # 헤더
        headers = ["버스ID", "출발지", "도착지", "출발시간", "도착시간", "소요시간(분)", "총 탑승인원"]
        ws.append(headers)
        header_row = ws.max_row
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border()

        # 데이터
        for i, route in enumerate(routes):
            vehicle = route.get('vehicle', {})
            row = [
                route.get('bus_id', ''),
                vehicle.get('start_location', ''),
                vehicle.get('end_location', vehicle.get('start_location', '')),
                route.get('departure_time', ''),
                route.get('arrival_time', ''),
                route.get('total_duration_min', 0),
                route.get('total_passengers', 0)
            ]
            ws.append(row)
            data_row = ws.max_row
            bus_color = BUS_COLORS[i % len(BUS_COLORS)]
            fill = PatternFill(start_color=bus_color, end_color=bus_color, fill_type="solid") if i % 2 == 0 else ALT_ROW_FILL
            for col_idx in range(1, len(row) + 1):
                cell = ws.cell(row=data_row, column=col_idx)
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border()
                # 버스ID 셀 색상
                if col_idx == 1:
                    cell.fill = PatternFill(start_color=bus_color, end_color=bus_color, fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                elif i % 2 == 0:
                    cell.fill = ALT_ROW_FILL

        # 컬럼 너비
        col_widths = [12, 20, 20, 12, 12, 14, 14]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _write_route_sheet(self, wb, routes):
        ws = wb.create_sheet("Route Detail")

        headers = ["버스ID", "순서", "탑승지", "주소", "탑승시간", "탑승인원", "구분"]
        ws.append(headers)
        header_row = ws.max_row
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border()

        for i, route in enumerate(routes):
            bus_color = BUS_COLORS[i % len(BUS_COLORS)]
            for stop in route.get('stops', []):
                if stop['type'] not in ('pickup', 'destination'):
                    continue
                row = [
                    route.get('bus_id', ''),
                    stop.get('order', '') + 1 if isinstance(stop.get('order'), int) else '',
                    stop.get('name', ''),
                    stop.get('address', '행사장'),
                    stop.get('pickup_time', ''),
                    stop.get('passenger_count', '') if stop['type'] == 'pickup' else '',
                    "탑승" if stop['type'] == 'pickup' else "도착"
                ]
                ws.append(row)
                data_row = ws.max_row
                for col_idx in range(1, len(row) + 1):
                    cell = ws.cell(row=data_row, column=col_idx)
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = thin_border()
                    if col_idx == 1:
                        cell.fill = PatternFill(start_color=bus_color, end_color=bus_color, fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif data_row % 2 == 0:
                        cell.fill = ALT_ROW_FILL

        col_widths = [12, 8, 20, 30, 12, 10, 8]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _write_passenger_sheet(self, wb, routes):
        ws = wb.create_sheet("Passenger")

        headers = ["이름", "버스ID", "탑승지", "탑승시간", "탑승인원"]
        ws.append(headers)
        header_row = ws.max_row
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border()

        for i, route in enumerate(routes):
            bus_color = BUS_COLORS[i % len(BUS_COLORS)]
            for stop in route.get('stops', []):
                if stop['type'] != 'pickup':
                    continue
                row = [
                    stop.get('name', ''),
                    route.get('bus_id', ''),
                    stop.get('address', ''),
                    stop.get('pickup_time', ''),
                    stop.get('passenger_count', 1)
                ]
                ws.append(row)
                data_row = ws.max_row
                for col_idx in range(1, len(row) + 1):
                    cell = ws.cell(row=data_row, column=col_idx)
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = thin_border()
                    if col_idx == 2:
                        cell.fill = PatternFill(start_color=bus_color, end_color=bus_color, fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif data_row % 2 == 0:
                        cell.fill = ALT_ROW_FILL

        col_widths = [20, 12, 35, 12, 10]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
